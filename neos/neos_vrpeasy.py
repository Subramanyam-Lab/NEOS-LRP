from lrp_easy_new import createLRP, write_to_txt_cvrplib_format
from solver_cvrp import dataCvrp
import os
import openpyxl
from dataparse import create_data
from datetime import datetime
import logging
import sys
from openpyxl import Workbook, load_workbook
from solver_cvrp_vrpsolver_modified_objective_instancepath import solve_instance


BFS = "/Users/waquarkaleem/NEOS-LRP-Codes-2/neos/dil_instances/BFS/f_rscc_hetero_ortools/1000"
phi_loc = '/Users/waquarkaleem/NEOS-LRP-Codes-2/pre_trained_model/f_rscc_hetero_ortools/model_phi_dnn_1000.onnx'
rho_loc = '/Users/waquarkaleem/NEOS-LRP-Codes-2/pre_trained_model/f_rscc_hetero_ortools/model_rho_dnn_1000.onnx'
existing_excel_file= "/Users/waquarkaleem/NEOS-LRP-Codes-2/results/Final_Results_Journal/f_rscc_hetero_ortools.xlsx" 
sheet_name = "f_rscc_hetero_ortools_1000"

log_dir = "log_files/mip_nn"
os.makedirs(log_dir, exist_ok=True)
# Directory containing the prodhon dataset
directory_path = "/Users/waquarkaleem/NEOS-LRP-Codes-2/prodhon_dataset"

try:
    workbook = load_workbook(existing_excel_file)
except FileNotFoundError:
    print(f"Excel file not found. Creating new file: {existing_excel_file}")
    workbook = Workbook()
    workbook.save(existing_excel_file)

if sheet_name not in workbook.sheetnames:
    workbook.create_sheet(sheet_name)
worksheet = workbook[sheet_name]

headings = [
    "Instance", "FLP", "VRP", "LRP(MIP+NN)", "NumRoutes_OptSol",
    "Exec time per depot(MIP+NN)", "initial solution generation time",
    "NN model execution time", "VRPSolverEasy computed VRP cost",
    "actual LRP cost(using VRPSolverEasy)",
    "avg solver_cvrp script execution time per depot",
    "total solver_cvrp script execution time",
    "VRPSolverEasy model solve time" , "BKS", "Optimization_gap_optsol", "Prediction_gap","min_num_routes_bfs", "Optimization_gap_bfs"
]

if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet.cell(1, 1).value is None:
    # Add headings to the first row
    for col, heading in enumerate(headings, start=1):
        worksheet.cell(row=1, column=col, value=heading)

def has_customers(instance_file_path):
    """Check if a VRP instance has any customers assigned."""
    with open(instance_file_path, 'r') as file:
        lines = file.readlines()
    try:
        demand_section_index = lines.index("DEMAND_SECTION\n")
        depot_section_index = lines.index("DEPOT_SECTION\n")
    except ValueError as e:
        raise ValueError(f"Section missing in file {instance_file_path}: {e}")
    demand_lines = lines[demand_section_index + 1:depot_section_index]
    # Remove depot demand (assumed to be the first line)
    customer_demands = demand_lines[1:]
    return len(customer_demands) > 0

def process_feasible_solutions(instance_subdir):
    # Initialize minimal cost to a large value
    minimal_cost = float('inf')
    # Initialize the number of routes associated with the minimal cost
    min_num_routes_bfs = 0
    
    # For each feasible solution folder in instance_subdir
    for solution_folder_name in os.listdir(instance_subdir):
        solution_folder_path = os.path.join(instance_subdir, solution_folder_name)
        if os.path.isdir(solution_folder_path) and solution_folder_name.startswith('feasible_solution_'):
            total_solver_cost = 0  # Total cost for the current feasible solution
            total_num_routes_bfs = 0  # Total number of routes for the current feasible solution
                
            # Loop over all VRP instances (depots) in the folder
            for filename in os.listdir(solution_folder_path):
                if filename.endswith(".txt"):
                    vrp_instance_file_path = os.path.join(solution_folder_path, filename)
                    try:
                        if not has_customers(vrp_instance_file_path):
                            # No customers assigned to this depot
                            solver_cost = 0.0
                            num_routes_bfs = 0
                            print(f"Instance: {filename} has zero customers.")
                        else:
                            # Solve the instance using the exact solver
                            solver_cost, num_routes_bfs, message, routes, solver_time = solve_instance(vrp_instance_file_path)
                            print(f"Instance: {filename}, Solver Cost: {solver_cost}, Number of Routes: {num_routes_bfs}")
                        total_solver_cost += solver_cost  # Sum up the costs
                        total_num_routes_bfs += num_routes_bfs  # Sum up the number of routes
                    except Exception as e:
                        print(f"Error processing instance {filename}: {e}")
                        continue
            # Update minimal cost and associated number of routes if this feasible solution has a lower total cost
            if total_solver_cost < minimal_cost:
                minimal_cost = total_solver_cost
                min_num_routes_bfs = total_num_routes_bfs
                print(f"New minimal cost found: {minimal_cost} with {min_num_routes_bfs} routes.")
    
    # Also process the final solution folder if it exists
    # final_solution_folder = os.path.join(instance_subdir, 'final_solution')
    # if os.path.exists(final_solution_folder):
    #     total_solver_cost = 0
    #     total_num_routes_bfs = 0
    #     for filename in os.listdir(final_solution_folder):
    #         if filename.endswith(".txt"):
    #             vrp_instance_file_path = os.path.join(final_solution_folder, filename)
    #             try:
    #                 if not has_customers(vrp_instance_file_path):
    #                     # No customers assigned to this depot
    #                     solver_cost = 0.0
    #                     num_routes_bfs = 0
    #                     print(f"Instance: {filename} has zero customers.")
    #                 else:
    #                     solver_cost, num_routes_bfs, message, routes, solver_time = solve_instance(vrp_instance_file_path)
    #                     print(f"Instance: {filename}, Solver Cost: {solver_cost}, Number of Routes: {num_routes_bfs}")
    #                 total_solver_cost += solver_cost
    #                 total_num_routes_bfs += num_routes_bfs
    #             except Exception as e:
    #                 print(f"Error processing instance {filename}: {e}")
    #                 continue
    #     # Update minimal cost and associated number of routes if final solution has a lower total cost
    #     if total_solver_cost < minimal_cost:
    #         minimal_cost = total_solver_cost
    #         min_num_routes_bfs = total_num_routes_bfs
    #         print(f"New minimal cost found in final solution: {minimal_cost} with {min_num_routes_bfs} routes.")
    
    return minimal_cost, min_num_routes_bfs

bks_dict = {
    "coord20-5-1.dat": 54793,
    "coord20-5-1b.dat": 39104,
    "coord20-5-2.dat": 48908,
    "coord20-5-2b.dat": 37542,
    "coord50-5-1.dat": 90111,
    "coord50-5-1b.dat": 63242,
    "coord50-5-2.dat": 88293,
    "coord50-5-2b.dat": 67308,
    "coord50-5-2bBIS.dat": 51822,
    "coord50-5-2BIS.dat": 84055,
    "coord50-5-3.dat": 86203,
    "coord50-5-3b.dat": 61830,
    "coord100-5-1.dat": 274814,
    "coord100-5-1b.dat": 213568,
    "coord100-5-2.dat": 193671,
    "coord100-5-2b.dat": 157095,
    "coord100-5-3.dat": 200079,
    "coord100-5-3b.dat": 152441,
    "coord100-10-1.dat": 287661,
    "coord100-10-1b.dat": 230989,
    "coord100-10-2.dat": 243590,
    "coord100-10-2b.dat": 203988,
    "coord100-10-3.dat": 250882,
    "coord100-10-3b.dat": 203114,
    "coord200-10-1.dat": 474850,
    "coord200-10-1b.dat": 375177,
    "coord200-10-2.dat": 448077,
    "coord200-10-2b.dat": 373696,
    "coord200-10-3.dat": 469433,
    "coord200-10-3b.dat": 362320
}


# bks_dict = {
#     "coord20-5-2.dat": 48908
# }

for filename in bks_dict.keys(): 
    file_path = os.path.join(directory_path, filename)
    if os.path.exists(file_path):
        print("Working on:", file_path)
    else:
        print("File not found:", file_path)
        break
    
    instance_subdir = os.path.join(BFS, os.path.splitext(filename)[0])
    os.makedirs(instance_subdir, exist_ok=True)

    # Call your function to get data
    # Create the log file name based on the input file's name
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  
    log_filename = f"{os.path.splitext(filename)[0]}_{current_time}.log"

    logging.basicConfig(
        level=logging.INFO,  # Set the logging level (INFO, WARNING, ERROR, etc.)
        format='%(asctime)s - %(levelname)s - %(message)s',  # Define the log message format
        handlers=[
            logging.FileHandler(os.path.join(log_dir, log_filename)),  # Log to a file
            logging.StreamHandler(sys.stdout)   # Log to stdout
        ]
    )
    logging.info(f'\n\n Working on file :{file_path}')
    ans = create_data(file_path)  # create_data is a function that processes the data
    
    
    lrp_st=datetime.now()
    lrp_solver = createLRP(ans)  #  createFlp is a function that creates the solver
    # lrp_result=lrp_solver.model(file_path,log_filename)
    lrp_result=lrp_solver.model(file_path, log_filename, instance_subdir,phi_loc, rho_loc)
    lrp_ed=datetime.now()

    # print(lrp_result)
    warmstart_time=lrp_result[4] #per depot
    nn_model_time=lrp_result[5]
    logging.info("\n\n")
    logging.info(f"Facility assignment decisions Results{lrp_result}")
    
    
    flp_dict={}
    for j in range(len(lrp_result[0])):
        if lrp_result[0][j]>0.5:
            ls=[]
            for i in range(len(lrp_result[1][j])):          
                if lrp_result[1][j][i]>0.5:
                    ls.append(i)
            flp_dict[j]=ls
    rout_dist={}
    fac_cust_dem={}
    cust_dem_fac={}
    for f in flp_dict:
        ls1=[]
        ls2=[]
        dem_sum=0
        for c in (flp_dict[f]):
            ls1.append(ans[3][c])
            dem_sum=dem_sum+ans[6][c]
            ls2.append(ans[6][c])
        ls1.insert(0,ans[2][f])
        ls2.insert(0,0)
        rout_dist[f]=ls1
        fac_cust_dem[f]=dem_sum
        cust_dem_fac[f]=ls2
    ass_result=[lrp_result[2],flp_dict,rout_dist,fac_cust_dem,cust_dem_fac]

    # for depot_id, customers in ass_result[1].items():
    #     depot_coords = [ass_result[2][depot_id][0]]  # First entry is depot coordinates
    #     customer_coords = ass_result[2][depot_id][1:]  # Customer coordinates (excluding depot)
    #     customer_demands = ass_result[4][depot_id][1:]  # Customer demands (excluding depot demand)
    #     vehicle_capacity = ans[4][0]  
    #     depot_customers = customer_coords  # Coordinates of customers only
    #     num_customers = len(customers)
    #     filename = f"cvrp_instance_depot_{os.path.basename(file_path).split('.')[0]}_depot_{depot_id}_customers_{num_customers}.txt"
    #     output_file_path = os.path.join(BFS, filename)

        # write_to_txt_cvrplib_format(depot_id, depot_customers, depot_coords, customer_demands, output_file_path, vehicle_capacity)

    ve_st=datetime.now()
    print("Running vrpeasy function")
    logging.info(f"The Input to vRP easy {ass_result}")

    print(f"The Input to vRP easy {ass_result}")
    logging.info(f"ANS (Data parsed from .dat file): {ans}")

    vrpeasy_solver=dataCvrp(ans,ass_result)
    vrp_easy_results=vrpeasy_solver.runVRPeasy()
    logging.debug("VRP Results",vrp_easy_results)
    ve_ed=datetime.now()

    od=len(flp_dict)
    
    lrp_exec=((lrp_ed-lrp_st).total_seconds())/od
    warmstart_time=warmstart_time/od
    tot_ve_exec=(ve_ed-ve_st).total_seconds()
    ve_exec=((ve_ed-ve_st).total_seconds())/od
    
    # Close the log file
    logging.shutdown()

    instance_name = os.path.basename(file_path)
    bks = bks_dict.get(instance_name)
    actual_lrp_cost_optsol = vrp_easy_results[0]
    
    vrp_routes_optsol = sum(vrp_easy_results[3])

    if bks is not None:
        gap = (abs(bks - actual_lrp_cost_optsol) / bks) * 100
        gap = round(gap, 2)  # Round to 2 decimal places
    else:
        bks = "N/A"
        gap = "N/A"

    if lrp_result[3] != 0:
        gap_vrp_perc = ((vrp_easy_results[1] - lrp_result[3]) / vrp_easy_results[1]) * 100
        gap_vrp_perc = abs(round(gap_vrp_perc, 2))  # Round to 2 decimal 
    else:
        gap_vrp_perc = "N/A"  

    if gap > 1.00:
        try:
            minimal_cost_bfs, min_num_routes_bfs = process_feasible_solutions(instance_subdir)
            actual_lrp_cost_bfssol = lrp_result[2] + minimal_cost_bfs
            if lrp_result[3] != 0:
                lrp_gap_bfs = ((bks - actual_lrp_cost_bfssol) / bks * 100)
                lrp_gap_bfs = abs(round(lrp_gap_bfs, 2))  # Round to 2 decimals
            else:
                lrp_gap_bfs = "N/A"
        except Exception as e:
            print(f"Error processing feasible solutions or calculating lrp_gap_bfs: {e}")
            minimal_cost_bfs = "NA"
            min_num_routes_bfs = "NA"
            lrp_gap_bfs = "N/A"
    else:
        minimal_cost_bfs = "NA"
        min_num_routes_bfs = "NA"
        lrp_gap_bfs = "N/A"  # Ensure lrp_gap_bfs is defined

    if lrp_gap_bfs != "N/A" and gap < lrp_gap_bfs:
        lrp_gap_bfs = gap

    # Create a new row
    new_row = [os.path.basename(file_path),lrp_result[2],lrp_result[3],lrp_result[2]+lrp_result[3], vrp_routes_optsol, lrp_exec, warmstart_time,nn_model_time,vrp_easy_results[1], vrp_easy_results[0],ve_exec,tot_ve_exec,vrp_easy_results[7], bks, gap, gap_vrp_perc, min_num_routes_bfs, lrp_gap_bfs]
        
    # Append the new row to the worksheet
    worksheet.append(new_row)

    # Save the modified Excel file
    workbook.save(existing_excel_file)